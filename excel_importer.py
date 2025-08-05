from openpyxl import load_workbook
from db import get_connection

def import_students_from_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    conn = get_connection()
    cursor = conn.cursor()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, roll, course = row

        # Check if roll number exists
        cursor.execute("SELECT COUNT(*) FROM Students WHERE roll_number = %s", (roll,))
        if cursor.fetchone()[0] == 0:
            cursor.execute(
                "INSERT INTO Students (student_name, roll_number, course_name) VALUES (%s, %s, %s)",
                (name, roll, course)
            )
        else:
            print(f"⚠️ Skipping duplicate roll number: {roll}")

    conn.commit()
    conn.close()
    print("✅ Student import completed.")
